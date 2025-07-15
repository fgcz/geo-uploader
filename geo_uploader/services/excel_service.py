import logging
import os
from collections import defaultdict

import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from geo_uploader.config import get_config
from geo_uploader.dto import SampleMetadata
from geo_uploader.services.file_service import FileService
from geo_uploader.utils.constants import (
    MD5_CHECKSUM_STARTROW,
    METADATA_PAIREDEND_STARTROW,
    METADATA_SAMPLES_STARTROW,
    SAMPLE_ATTRIBUTE_COLUMN_MAPPINGS,
    SAMPLE_NAME_TO_COLUMNS,
)
from geo_uploader.utils.metadata.edit_metadata import (
    insert_column,
    insert_row,
    insert_sample_rows,
    reapply_hidden_dropdown,
    remove_column,
    remove_row,
)

logger = logging.getLogger(__name__)

# no header
# used to know from where to start saving the spreadsheet
# used to know from where relatively to insert/delete rows related to study
# metadata_study_startrow = 11 (OLD)
metadata_study_startrow = 12

# including header
# metadata_samples_startrow = 33 (OLD)
metadata_samples_startrow = 38

metadata_protocols_instructions = 54
# no header
# metadata_protocols_startrow = 56 (OLD)
metadata_protocols_startrow = 57

# including header
# metadata_pairedend_startrow = 75 (OLD)
metadata_pairedend_startrow = 76

# where to insert a new column when user wants to add more sample data
metadata_samples_column_insert = (
    6  # starting from six on you can insert a column, but until when?
)
metadata_samples_column_raw = 17

metadata_samples_column = 20

md5_checksum_startrow = 9


class ExcelService:
    def __init__(self, config=None, logger=None, file_service=None):
        self.config = config or get_config()
        self.logger = logger or current_app.logger
        self.file_service = file_service or FileService()

    def copy_new_session_metadata(self, session_folder_path):
        """Creates a new metadata Excel file into the session_folder_path"""
        wb = load_workbook(self.config.BASE_EXCEL)
        excel_path = os.path.join(session_folder_path, "Metadata.xlsx")
        wb.save(excel_path)

    @classmethod
    def autocomplete_metadata(
        cls,
        samples: list[SampleMetadata],
        protocols_displacement: int,
        sheet: Worksheet,
        is_singlecell: bool,
    ):
        """
        Complete metadata for each sample in the sheet.

        Args:
            samples: List of sample metadata objects
            protocols_displacement: Displacement for protocols section
            sheet: The Worksheet to update
            is_singlecell: Whether this is single cell data (as opposed to bulk)
        """
        cls._autocomplete_bulk_metadata(samples, sheet)
        cls._autocomplete_bulk_pairedend_section(
            samples, sheet, protocols_displacement
        )

    @staticmethod
    def autocomplete_md5checksums(md5tsv, sheet):
        """
        Read the md5tsv path with filename, file_type, and md5sum columns,
        and complete the sheet with the file:md5 for each file.
        Position where to write is hardcoded.
        """

        # Read the new file format
        df = pd.read_csv(md5tsv, sep="\t")

        # Separate raw and processed file entries
        raw_files = df[df["file_type"] == "raw"][["file_name", "md5sum"]]
        processed_files = df[df["file_type"] == "processed"][["file_name", "md5sum"]]

        # Convert to iterable tuples
        raw_tuples = list(raw_files.itertuples(index=False, name=None))
        processed_tuples = list(processed_files.itertuples(index=False, name=None))

        # Populate sheet for raw files
        for current_pos, (raw_filename, raw_md5) in enumerate(raw_tuples):
            sheet.cell(MD5_CHECKSUM_STARTROW + current_pos, 1).value = raw_filename
            sheet.cell(MD5_CHECKSUM_STARTROW + current_pos, 2).value = raw_md5

        # Populate sheet for processed files
        for current_pos, (processed_filename, processed_md5) in enumerate(
            processed_tuples
        ):
            sheet.cell(
                MD5_CHECKSUM_STARTROW + current_pos, 6
            ).value = processed_filename
            sheet.cell(MD5_CHECKSUM_STARTROW + current_pos, 7).value = processed_md5

    @classmethod
    def _autocomplete_bulk_metadata(
        cls, samples: list[SampleMetadata], sheet: Worksheet
    ):
        start_row = METADATA_SAMPLES_STARTROW + 1

        for idx, sample in enumerate(samples):
            sample_row = start_row + idx

            # Set basic metadata values -> name, organism, instrument_model
            cls._set_basic_sample_metadata(sample, sample_row, sheet)

            # Map the raw files to their respective columns
            for i, file_info in enumerate(sample.raw_file_paths):
                raw_column_number = SAMPLE_NAME_TO_COLUMNS["raw_file_1"] + i
                sheet.cell(sample_row, raw_column_number).value = file_info.file_name

            # Handle processed files
            cls._set_bulk_processed_files(sample, sample_row, sheet)

            # Set paired-end status
            is_paired = len(sample.raw_file_paths) > 1
            pairedend_value = "paired-end" if is_paired else "single"
            sheet.cell(
                sample_row, SAMPLE_NAME_TO_COLUMNS["single_or_pairedend"]
            ).value = pairedend_value

    @classmethod
    def _set_bulk_processed_files(
        cls, sample: SampleMetadata, row: int, sheet: Worksheet
    ) -> None:
        """Set processed file information in the sheet"""
        processed_files = sample.processed_file_paths

        # Map the processed files to their respective columns
        for i, file_info in enumerate(processed_files, start=1):
            column_name = f"processed_file_{i}"
            if column_name in SAMPLE_NAME_TO_COLUMNS:
                column = SAMPLE_NAME_TO_COLUMNS[column_name]
                sheet.cell(row, column).value = file_info.file_name

    @classmethod
    def _set_basic_sample_metadata(
        cls, sample: SampleMetadata, row: int, sheet: Worksheet
    ):
        """Set basic metadata like name, organism, instrument model"""
        direct_mappings = SAMPLE_ATTRIBUTE_COLUMN_MAPPINGS

        for attr_name, column_name in direct_mappings.items():
            value = getattr(sample, attr_name, None)
            if value:
                col_idx = SAMPLE_NAME_TO_COLUMNS.get(column_name)
                if col_idx:
                    sheet.cell(row, col_idx).value = value

    @classmethod
    def _autocomplete_bulk_pairedend_section(
        cls,
        samples: list[SampleMetadata],
        sheet: Worksheet,
        protocols_displacement: int,
    ):
        start_row = METADATA_PAIREDEND_STARTROW + 1 + protocols_displacement
        pairedend_samples = [
            sample for sample in samples if len(sample.raw_file_paths) > 1
        ]
        for idx, sample in enumerate(pairedend_samples):
            sample_row = start_row + idx

            sheet.cell(sample_row, 1).value = sample.raw_file_paths[0].file_name
            sheet.cell(sample_row, 2).value = sample.raw_file_paths[1].file_name

            # Handle R3 and R4 or I1 and I2 if available

    @staticmethod
    def _extract_sample_info(filename: str):
        """Extract sample name, sample number, and lane from a filename"""
        from re import search

        match = search(r"(.+?)_S(\d+)_L(\d+)", filename)
        if match:
            name = match.group(1)  # Part before _S (e.g., 321081_2-running_1)
            sample_num = match.group(2)  # Sample number (e.g., S3)
            lane = match.group(3)  # Lane number (e.g., L001 or L002)
            return name, sample_num, lane
        return None, None, None

    def prepare_open_metadata(self, session_title, specific_sheet="Metadata"):
        excel_path = self.file_service.get_session_folderpath(
            session_title, "Metadata.xlsx"
        )

        try:
            wb = load_workbook(excel_path)
        except Exception as e:
            logger.warning(f"Exception reading the excel workbook {e}")
            return

        sheet = wb[specific_sheet]  # select by name
        return wb, sheet

    def prepare_close_metadata(self, wb, session_title):
        excel_path = self.file_service.get_session_folderpath(
            session_title, "Metadata.xlsx"
        )
        try:
            wb.save(excel_path)
        except Exception as e:
            logger.warning(f"Exception reading the excel workbook {e}")


def save_add_contributor(
    metadata_study_length, metadata_supplementary_number, excel_path
):
    """inserts a new contributor row to the spreadsheet"""

    insert_position = metadata_study_startrow + metadata_study_length - 1
    insert_position = insert_position - metadata_supplementary_number
    insert_row(
        excel_path, "Metadata", insert_row=insert_position, cell_type="contributor"
    )


def save_remove_contributor(
    metadata_study_length, metadata_supplementary_number, excel_path
):
    """removes a contributor row from the spreadsheet"""

    delete_row = metadata_study_startrow + metadata_study_length
    delete_row = delete_row - metadata_supplementary_number - 1
    remove_row(excel_path, "Metadata", delete_row=delete_row)


def save_add_supplementaryfile(metadata_study_length, excel_path):
    """inserts a new supplementary_file row to the spreadsheet"""

    insert_position = metadata_study_startrow + metadata_study_length
    insert_row(
        excel_path, "Metadata", insert_row=insert_position, cell_type="supplementary"
    )


def save_remove_supplementaryfile(metadata_study_length, excel_path):
    """removes a supplementary_file row from the spreadsheet"""

    delete_row = metadata_study_startrow + metadata_study_length - 1
    remove_row(excel_path, "Metadata", delete_row=delete_row)


def save_add_step(_session, excel_path):
    """inserts a new data processing step row to the spreadsheet"""

    insert_position = (
        metadata_protocols_startrow + _session.metadata_protocol_displacement
    )
    insert_position += _session.metadata_protocol_length - 1
    insert_position -= (
        _session.metadata_processedfiles_number + 1
    )  # for *genome build/assembly
    insert_row(excel_path, "Metadata", insert_row=insert_position, cell_type="step")


def save_remove_step(_session, excel_path):
    """removes a data processing step row from the spreadsheet"""

    delete_row = metadata_protocols_startrow + _session.metadata_protocol_displacement
    delete_row += _session.metadata_protocol_length - 1
    delete_row -= (
        _session.metadata_processedfiles_number + 1
    )  # for *genome build/assembly
    remove_row(excel_path, "Metadata", delete_row=delete_row)


def save_add_format(_session, excel_path):
    """inserts a data format row to the spreadsheet"""

    insert_position = (
        metadata_protocols_startrow + _session.metadata_protocol_displacement
    )
    insert_position += _session.metadata_protocol_length - 1
    insert_row(excel_path, "Metadata", insert_row=insert_position, cell_type="format")


def save_remove_format(_session, excel_path):
    """removes a data format row from the spreadsheet"""

    delete_row = metadata_protocols_startrow + _session.metadata_protocol_displacement
    delete_row += _session.metadata_protocol_length - 1

    remove_row(excel_path, "Metadata", delete_row=delete_row)


def resize_sample_columns(excel_path, samples, previous_sample_width):
    """Add or remove necessary columns to the samples section"""

    new_width = len(samples[0])
    size_change = new_width - previous_sample_width
    if size_change > 0:
        for _i in range(abs(size_change)):
            # we insert a column for the place, then the column content will be overwritten from scratch
            # make sure not to touch columns <4, because of paired end section
            insert_column(
                excel_path,
                "Metadata",
                insert_column=metadata_samples_column_insert,
                header_line=metadata_samples_startrow,
            )
    elif size_change < 0:
        for _i in range(abs(size_change)):
            # we remove a random column, then the column content will be overwritten from scratch
            # make sure not to touch columns <4, because of paired end section
            remove_column(
                excel_path, "Metadata", delete_column=metadata_samples_column_insert
            )


def save_study_metadata(study, sheet):
    """Populates the study section from the study data.
    Rewrites only column 1/2, and same rows as study data"""

    start_row = metadata_study_startrow

    for row in range(start_row, start_row + len(study)):
        sheet.cell(row, 1).value = study[row - start_row][0]
        sheet.cell(row, 2).value = study[row - start_row][1]


def save_sample_metadata(samples, sample_displacement, sheet):
    """Populate the sample section from the samples data.
    The width and height are taken from the samples data."""

    new_width = len(samples[0])
    start_row = metadata_samples_startrow + sample_displacement

    for sample_row in range(start_row, start_row + len(samples)):
        for sample_col in range(1, new_width + 1):
            sheet.cell(sample_row, sample_col).value = samples[sample_row - start_row][
                sample_col - 1
            ]


def save_protocol_metadata(protocol, protocol_displacement, sheet):
    """Populate the protocol section from the protocol data.
    Rewrites only column 1/2, and same rows as protocol data"""

    new_protocol_startrow = metadata_protocols_startrow + protocol_displacement

    for row in range(new_protocol_startrow, new_protocol_startrow + len(protocol)):
        sheet.cell(row, 1).value = protocol[row - new_protocol_startrow][0]
        sheet.cell(row, 2).value = protocol[row - new_protocol_startrow][1]


def resize_samples(open_path, is_singlecell, sample_length, raw_length):
    """
    Add rows in case the sample_length overlaps with the other sections\n
    Add columns in case of SingleCell: 1 column for processed, raw_length - 4(already by default) for raw
    """

    reapply_hidden_dropdown(open_path, "Metadata")

    # remove the paired/single-end in case of bulk
    samples_length = sample_length - 1 if not is_singlecell else sample_length

    protocols_displacement = 0
    # we add rows whether it is single cell or bulk
    if (
        metadata_samples_startrow + samples_length + 1
        >= metadata_protocols_instructions
    ):
        # have at least 1 extra row in case they are equal, or 1 + overlap
        rows_extra = (
            1
            + (metadata_samples_startrow + samples_length + 1)
            - metadata_protocols_instructions
        )

        # the row position where to insert the rows is hard coded in the insert_sample_rows function
        insert_sample_rows(
            open_path, "Metadata", metadata_samples_startrow + 1, rows_extra
        )

        protocols_displacement = rows_extra

    if is_singlecell:
        # only single cell needs to add columns to samples

        # insert 1 processed column always
        # in case we are uploading only raw files, we would have one extra column, but whatever
        insert_column(
            open_path,
            "Metadata",
            metadata_samples_column_raw,
            header_line=metadata_samples_startrow,
            file_column=True,
        )

        # insert additional raw columns
        if raw_length > 4:
            raw_columns_to_insert = raw_length - 4
            for _i in range(raw_columns_to_insert):
                insert_column(
                    open_path,
                    "Metadata",
                    metadata_samples_column_raw + 3,
                    header_line=metadata_samples_startrow,
                    file_column=True,
                )

    return protocols_displacement


def load_dropdowns(sheet):
    dropdown_instrument = [sheet[f"A{i}"].value for i in range(2, 73)]
    dropdown_molecule = [
        "polyA RNA",
        "total RNA",
        "nuclear RNA",
        "cytoplasmic RNA",
        "genomic DNA",
        "protein",
        "other",
    ]
    dropdown_library = [sheet[f"B{i}"].value for i in range(2, 49)]

    return dropdown_molecule, dropdown_instrument, dropdown_library


def load_metadata(sheet, _session):
    """take data from the sheet into an object representable in the handsontable. \n
    The dimensions of the hands on table need to be exact to the metadata, because when we save the changes we take
    the dimensions from the handsontable."""

    # ----------- STUDY FORM --------------
    study_list_data = []
    current_cell = metadata_study_startrow
    while sheet[f"A{current_cell}"].value is not None:
        study_list_data.append(
            [sheet[f"A{current_cell}"].value, sheet[f"B{current_cell}"].value]
        )
        current_cell += 1

    # ----------- SAMPLE FORM --------------
    samples_list_data = []

    # account for header
    sample_start = metadata_samples_startrow + _session.metadata_samples_displacement
    for i in range(sample_start, sample_start + _session.metadata_samples_length + 1):
        current_sample = []
        for j in range(1, _session.metadata_samples_width + 1):
            current_sample.append(sheet.cell(i, j).value)
        samples_list_data.append(current_sample)

    # ----------- PROTOCOL FORM --------------
    protocol_list_data = []
    protocol_start = (
        metadata_protocols_startrow + _session.metadata_protocol_displacement
    )
    for i in range(
        protocol_start, protocol_start + _session.metadata_protocol_length
    ):  # no header
        protocol_list_data.append([sheet[f"A{i}"].value, sheet[f"B{i}"].value])

    # ----------- PAIREDEND FORM ----------
    pairedend_list_data = []

    # +1 to account for the header
    pairedend_start = (
        metadata_pairedend_startrow + _session.metadata_pairedend_displacement
    )
    i = pairedend_start
    # for i in range(pairedend_start, pairedend_start + _session.metadata_samples_length + 1):
    while sheet[f"A{i}"].value is not None:
        # Todo, calculate for more columns if R3, R4
        pairedend_list_data.append(
            [
                sheet[f"A{i}"].value,
                sheet[f"B{i}"].value,
                sheet[f"C{i}"].value,
                sheet[f"D{i}"].value,
            ]
        )
        i += 1

    list_data = []
    for i in range(1, 180):
        current_line = []
        for j in range(1, _session.metadata_samples_width + 1):
            current_line.append(sheet.cell(i, j).value)
        list_data.append(current_line)

    return_variables = {
        "list_data": list_data,
        "study_list_data": study_list_data,
        "samples_list_data": samples_list_data,
        "protocol_list_data": protocol_list_data,
        "pairedend_list_data": pairedend_list_data,
    }
    return return_variables
