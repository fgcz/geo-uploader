from copy import copy

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formatting.rule import Rule
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def copy_cell_content(
    source_ws, target_ws, insert_row, rows_to_skip, insert_col, cols_to_skip
):
    """Works with value, data_type, _style, _hyperlink, comment
    data_type is 's', 'n'
    total number of comments in source is 43
    _hyperlink is empty
    """
    for (row, col), source_cell in source_ws._cells.items():
        if row == insert_row and rows_to_skip == -1:
            # this row will not be passed on
            continue

        target_row = (
            row + rows_to_skip if row >= insert_row else row
        )  # Keep rows above insert_row unchanged
        target_col = col + cols_to_skip if col >= insert_col else col

        target_cell = target_ws.cell(column=target_col, row=target_row)

        # Copy cell content and type
        target_cell.value = source_cell.value
        target_cell.data_type = source_cell.data_type

        # Copy cell style
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        # Copy hyperlink
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        # Copy comment
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

    # Next, apply the same style to the skipped rows
    for row in range(insert_row, insert_row + rows_to_skip):
        for col in range(
            1, source_ws.max_column + 1
        ):  # Apply to all columns in the row
            source_cell = source_ws.cell(row=insert_row, column=col)
            target_cell = target_ws.cell(row=row, column=col)

            # Copy the style from the source row that is being shifted down
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

    # Adjust the loop to copy styles from right to left when removing columns
    if cols_to_skip < 0:
        range_start = insert_col + cols_to_skip - 1
        range_end = insert_col - 1
        step = -1
    else:
        range_start = insert_col
        range_end = insert_col + cols_to_skip
        step = 1

    # Next, apply the same style to the skipped column
    for col in range(range_start, range_end, step):
        for row in range(1, source_ws.max_row):  # Apply to all rows in the column
            source_cell = source_ws.cell(row=row, column=insert_col - 1)
            target_cell = target_ws.cell(row=row, column=col)

            # Copy the style from the source row that is being shifted down
            target_cell._style = copy(source_cell._style)
            target_cell.comment = copy(source_cell.comment)


def copy_data_validators(
    source_ws, target_ws, insert_row, rows_to_skip, insert_col, cols_to_skip
):
    target_ws.data_validations = copy(source_ws.data_validations)

    # Check and see if you can retrieve this data from the hidden sheet.

    for target_dv in target_ws.data_validations.dataValidation:
        new_ranges = []
        for cell_range in list(target_dv.sqref):
            new_range = copy(cell_range)
            if insert_row < cell_range.min_row:
                min_row_shifted = cell_range.min_row + rows_to_skip
                max_row_shifted = cell_range.max_row + rows_to_skip
                if min_row_shifted <= 0:
                    raise ValueError(f"Invalid shift value: row_shift={rows_to_skip}")

                new_range.min_row = min_row_shifted
                new_range.max_row = min(max_row_shifted, 1048576)  # Excel max row limit
                # pass
            elif insert_row <= cell_range.max_row and insert_row >= cell_range.min_row:
                if cell_range.max_row == cell_range.min_row and rows_to_skip == -1:
                    # a single cell validator
                    # validator to be removed
                    continue
                else:
                    # a range, and we want to slide the max row
                    max_row_shifted = cell_range.max_row + rows_to_skip
                    new_range.max_row = min(max_row_shifted, 1048576)
            else:
                # insert_row > cell_range.max_row, so it says as it is
                pass

            if insert_col <= cell_range.min_col:
                min_col_shifted = cell_range.min_col + cols_to_skip
                max_col_shifted = cell_range.max_col + cols_to_skip
                if min_col_shifted <= 0:
                    raise ValueError(f"Invalid shift value: col_shift={cols_to_skip}")

                new_range.min_col = min_col_shifted
                new_range.max_col = min(
                    max_col_shifted, 16384
                )  # Excel max column limit
            elif insert_col <= cell_range.max_col and insert_col >= cell_range.min_col:
                if cell_range.max_col == cell_range.min_col and cols_to_skip == -1:
                    # A single cell validator in a column, so we remove it
                    continue
                elif cell_range.max_col == cell_range.min_col and cols_to_skip == 1:
                    # we slide both min and max
                    new_range.min_col = cell_range.min_col + cols_to_skip
                    new_range.max_col = cell_range.max_col + cols_to_skip

                else:
                    # A range, and we want to slide the max column
                    max_col_shifted = cell_range.max_col + cols_to_skip
                    new_range.max_col = min(max_col_shifted, 16384)
            else:
                # insert_col > cell_range.max_col, so it stays as it is
                pass

            new_ranges.append(new_range)
        # Clear the original sqref
        target_dv.sqref = set()
        # Add all the new ranges to sqref
        for new_range in new_ranges:
            target_dv.sqref.add(new_range)


def copy_row_and_column_styles(
    source_ws, target_ws, insert_row, rows_to_skip, insert_col, cols_to_skip
):
    """Copy row and column dimensions/styles from the source sheet to the target sheet,
    starting from insert_row and skipping a specified number of rows."""

    # Copy column dimensions and apply standard style for inserted columns

    for col_key, dim in source_ws.column_dimensions.items():
        col_index = column_index_from_string(col_key)

        if col_index >= insert_col:
            new_col_letter = get_column_letter(col_index + cols_to_skip)
            target_ws.column_dimensions[new_col_letter] = copy(dim)
        else:
            target_ws.column_dimensions[col_key] = copy(dim)

    # Copy row dimensions and apply standard style for inserted rows
    for row_key, dim in source_ws.row_dimensions.items():
        # only gives the row dimensions which have a style different, so the inserted rows will have a default style of standard

        if row_key >= insert_row:
            target_ws.row_dimensions[row_key + rows_to_skip] = copy(
                dim
            )  # Copy the original row style
        else:
            # Copy original row dimensions for rows above the insert_row
            target_ws.row_dimensions[row_key] = copy(dim)


def extract_conditional_formatting(sheet):
    """Extract conditional formatting rules from the sheet."""
    return list(sheet.conditional_formatting)


def shift_conditional_formatting(
    cf_rules, insert_row, rows_to_skip, insert_col, cols_to_skip
):
    shifted_rules = []
    for cf in cf_rules:
        # Shift all the ranges within the conditional formatting object
        shifted_sqref = []

        for cell_range in cf.sqref:
            if insert_row < cell_range.min_row:
                min_row_shifted = cell_range.min_row + rows_to_skip
                max_row_shifted = cell_range.max_row + rows_to_skip
                if min_row_shifted <= 0:
                    raise ValueError(f"Invalid shift value: row_shift={rows_to_skip}")

                cell_range.min_row = min_row_shifted
                cell_range.max_row = min(
                    max_row_shifted, 1048576
                )  # Excel max row limit
                # pass
            elif insert_row <= cell_range.max_row and insert_row >= cell_range.min_row:
                # If the insert_row is within the range, we adjust the end row
                # This means we shift the max_row only
                max_row_shifted = cell_range.max_row + rows_to_skip
                cell_range.max_row = min(max_row_shifted, 1048576)
            else:
                # insert_row > cell_range.max_row, so it says as it is
                pass

            if insert_col < cell_range.min_col:
                min_col_shifted = cell_range.min_col + cols_to_skip
                max_col_shifted = cell_range.max_col + cols_to_skip
                if min_col_shifted <= 0:
                    raise ValueError(
                        f"Invalid shift value: cols_to_skip={cols_to_skip}"
                    )
                cell_range.min_col = min_col_shifted
                cell_range.max_col = min(
                    max_col_shifted, 16384
                )  # Excel max column limit

            elif insert_col >= cell_range.min_col and insert_col <= cell_range.max_col:
                # If the insert_col is within the range, we adjust the end column
                # This means we shift the max_col only
                if cell_range.max_col == cell_range.min_col:
                    min_col_shifted = cell_range.min_col + cols_to_skip
                    cell_range.min_col = min(min_col_shifted, 16384)
                max_col_shifted = cell_range.max_col + cols_to_skip
                cell_range.max_col = min(max_col_shifted, 16384)
            else:
                # insert_col > cell_range.max_col, so it says as it is
                pass

            shifted_sqref.append(cell_range)

        # Create a new ConditionalFormatting object with shifted ranges and original rules
        shifted_cf = ConditionalFormatting(sqref=shifted_sqref, cfRule=cf.cfRule)
        shifted_rules.append(shifted_cf)
    return shifted_rules


def apply_conditional_formatting(sheet, cf_rules):
    for cf in cf_rules:
        for rule in cf.rules:
            if isinstance(rule, Rule):
                sheet.conditional_formatting.add(str(cf.sqref), rule)
            else:
                raise ValueError(
                    "Only instances of openpyxl.formatting.rule.Rule may be added"
                )


def copy_conditional_formatting_with_shift(
    src_sheet, tgt_sheet, insert_row, rows_to_skip, insert_col, cols_to_skip
):
    """Copy and shift conditional formatting rules from source to target sheet."""

    cf_rules = extract_conditional_formatting(src_sheet)
    shifted_cf_rules = shift_conditional_formatting(
        cf_rules, insert_row, rows_to_skip, insert_col, cols_to_skip
    )
    apply_conditional_formatting(tgt_sheet, shifted_cf_rules)


def reapply_hidden_dropdown(open_path, sheet_title):
    wb = load_workbook(open_path)
    ws = wb[sheet_title]

    dv = DataValidation(
        type="list",
        formula1="='Data validation'!$A$2:$A$73",  # Source of the dropdown list
        showErrorMessage=True,
        showInputMessage=True,
        showDropDown=False,  # Show dropdown arrow
        allowBlank=True,  # Allow blank entries
        promptTitle="",
        errorStyle="stop",
        error="Select a value from the drop down menu",
        prompt="Click on arrowhead.  To view complete list, use scrollbar or up/down arrows on keyboard (increasing window size or zoom level may improve scroll bar function).",
        errorTitle="",
    )
    # todo this is hardcoded
    dv.sqref = "M39:M53"
    ws.add_data_validation(dv)

    dv = DataValidation(
        type="list",
        formula1="='Data validation'!$B$2:$B$49",  # Source of the dropdown list
        showErrorMessage=True,
        showInputMessage=True,
        showDropDown=False,  # Show dropdown arrow
        allowBlank=True,  # Allow blank entries
        promptTitle="",
        errorStyle="stop",
        error="Select a value from the drop down menu",
        prompt="Click on arrowhead.  To view complete list, use scrollbar or up/down arrows on keyboard (increasing window size or zoom level may improve scroll bar function).",
        errorTitle="",
    )
    # todo this is hardcoded
    dv.sqref = "C39:C53"
    ws.add_data_validation(dv)

    wb.save(open_path)


def insert_sample_rows(open_path, sheet_title, insert_row, rows_to_skip=0):
    wb = load_workbook(open_path)
    ws = wb[sheet_title]  # select by name
    ws2 = wb.create_sheet("TemporaryPlaceholder", 0)
    ws2.sheet_properties.tabColor = copy(ws.sheet_properties.tabColor)

    insert_col = 0
    columns_to_skip = 0
    copy_cell_content(ws, ws2, insert_row, rows_to_skip, insert_col, columns_to_skip)
    copy_data_validators(ws, ws2, insert_row, rows_to_skip, insert_col, columns_to_skip)
    copy_row_and_column_styles(
        ws, ws2, insert_row, rows_to_skip, insert_col, columns_to_skip
    )
    copy_conditional_formatting_with_shift(
        ws, ws2, insert_row, rows_to_skip, insert_col, columns_to_skip
    )

    # # Remove the original sheet, rename the new sheet, and save the workbook
    wb.remove(ws)
    ws2.title = sheet_title
    wb.save(open_path)
    # wb.save("./TestInsertColumn5.xlsx")


def remove_column(open_path, sheet_title, delete_column):
    wb = load_workbook(open_path)
    ws = wb[sheet_title]  # select by name
    ws2 = wb.create_sheet("TemporaryPlaceholder", 0)
    ws2.sheet_properties.tabColor = copy(ws.sheet_properties.tabColor)

    columns_to_skip = -1
    delete_column += 1
    copy_cell_content(ws, ws2, 0, 0, delete_column, columns_to_skip)
    copy_data_validators(ws, ws2, 0, 0, delete_column - 1, columns_to_skip)
    copy_row_and_column_styles(ws, ws2, 0, 0, delete_column, columns_to_skip)
    copy_conditional_formatting_with_shift(
        ws, ws2, 0, 0, delete_column, columns_to_skip
    )

    wb.remove(ws)
    ws2.title = sheet_title
    wb.save(open_path)
    # wb.save("./TestInsertColumn5.xlsx")
    # wb.save("./Metadata.xlsx")


def insert_column(
    open_path, sheet_title, insert_column, header_line, file_column=False
):
    # By design, we can't insert/remove the first 4 columns due paired-end table
    # we can only insert on the right of **tisue, which is column 4, so we can insert 5
    # when we insert 5, means we take the comment of tissue
    # Adding into column 5, we take the coments from column 4

    wb = load_workbook(open_path)
    ws = wb[sheet_title]  # select by name
    ws2 = wb.create_sheet("TemporaryPlaceholder", 0)
    ws2.sheet_properties.tabColor = copy(ws.sheet_properties.tabColor)

    columns_to_skip = 1
    copy_cell_content(ws, ws2, 0, 0, insert_column, columns_to_skip)
    if file_column:
        source_cell = ws2.cell(row=header_line, column=insert_column - 1)
        target_cell = ws2.cell(row=header_line, column=insert_column)
        target_cell.value = source_cell.value
    copy_data_validators(ws, ws2, 0, 0, insert_column, columns_to_skip)
    copy_row_and_column_styles(ws, ws2, 0, 0, insert_column, columns_to_skip)
    copy_conditional_formatting_with_shift(
        ws, ws2, 0, 0, insert_column, columns_to_skip
    )

    wb.remove(ws)
    ws2.title = sheet_title
    wb.save(open_path)
    # wb.save("./TestInsertColumn5.xlsx")


def insert_row(open_path, sheet_title, insert_row, cell_type):
    """Insert an empty row in the current position, and copy style of cell above"""
    wb = load_workbook(open_path)
    ws = wb[sheet_title]  # select by name
    ws2 = wb.create_sheet("TemporaryPlaceholder", 0)
    ws2.sheet_properties.tabColor = copy(ws.sheet_properties.tabColor)
    keep_validators = True
    # print(ws.max_column)

    copy_cell_content(ws, ws2, insert_row, 1, 0, 0)

    target_cell = ws2.cell(row=insert_row, column=1)
    if cell_type == "contributor" or cell_type == "step" or cell_type == "format":
        source_cell = ws2.cell(row=insert_row + 1, column=1)
        # inserting at a contributor cell position
        target_cell.value = source_cell.value
        if source_cell.value == "*data processing step":
            source_cell.value = "data processing step"
            target_cell.value = "*data processing step"

        if source_cell.value == "*processed data files format and content":
            source_cell.value = "processed data files format and content"
            target_cell.value = "*processed data files format and content"
        target_cell.comment = copy(source_cell.comment)
        # Switch column 2 content so it looks as if we inserted down
        (
            ws2.cell(row=insert_row + 1, column=2).value,
            ws2.cell(row=insert_row, column=2).value,
        ) = (
            ws2.cell(row=insert_row, column=2).value,
            ws2.cell(row=insert_row + 1, column=2).value,
        )

    if cell_type == "supplementary":
        previous_cell = ws2.cell(row=insert_row - 1, column=1)
        # inserting at an empty cell
        # not guaranteed to find it from previous cell
        target_cell.value = "supplementary file"
        target_cell._style = copy(previous_cell._style)

        if previous_cell.value == "supplementary file":
            # find and extend the validator
            copy_data_validators(ws, ws2, insert_row - 1, 1, 0, 0)
            keep_validators = False
        else:
            # create the validator from the beginning
            dv = DataValidation(
                showInputMessage=True,
                promptTitle="",
                error="Select a value from the drop down menu",
                prompt="List the name of any processed data files (one per row) that were derived from multiple samples. For instance, bulkRNA-seq tables that include library names as headers, or 'merged' peak files.",
                errorStyle="stop",
            )
            dv.sqref = f"B{insert_row}"
            ws.add_data_validation(dv)
            copy_data_validators(ws, ws2, insert_row + 1, 1, 0, 0)
            keep_validators = False

    if keep_validators:
        copy_data_validators(ws, ws2, insert_row, 1, 0, 0)

    # if user removes a supply and then adds it again, the dv will be coppied down to supplementary file. But it needs to have a different dv.
    copy_row_and_column_styles(ws, ws2, insert_row, 1, 0, 0)
    copy_conditional_formatting_with_shift(ws, ws2, insert_row, 1, 0, 0)

    # # Remove the original sheet, rename the new sheet, and save the workbook
    wb.remove(ws)
    ws2.title = sheet_title
    wb.save(open_path)
    # wb.save("./TestInsertColumn5.xlsx")


def remove_row(open_path, sheet_title, delete_row):
    wb = load_workbook(open_path)
    ws = wb[sheet_title]  # select by name
    ws2 = wb.create_sheet("TemporaryPlaceholder", 0)
    ws2.sheet_properties.tabColor = copy(ws.sheet_properties.tabColor)

    copy_cell_content(ws, ws2, delete_row, -1, 0, 0)
    copy_data_validators(ws, ws2, delete_row, -1, 0, 0)
    copy_row_and_column_styles(ws, ws2, delete_row, -1, 0, 0)
    copy_conditional_formatting_with_shift(ws, ws2, delete_row, -1, 0, 0)

    wb.remove(ws)
    ws2.title = sheet_title
    wb.save(open_path)
    # wb.save("./TestInsertColumn5.xlsx")
    # wb.save("./Metadata.xlsx")


# REMEMBER TO REMOVE THE wb.save("./Metadata.xlsx") from insert column


# open_path = "./seq_template.xlsx"
# open_path2 = "./TestInsertColumn5.xlsx"
# wb = load_workbook(open_path)
# ws = wb["Data validation"]

# for row in ws.iter_rows(min_row=1, max_row=74, min_col=1, max_col=3):
#     print([cell.value for cell in row])

# insert_column(open_path, "Metadata", 6)
# remove_column(open_path, "Metadata", 7)
# insert_sample_rows(open_path, "Metadata", 1)
# insert_row(open_path, "Metadata", 15, "")
# insert_row(open_path, "Metadata", 21, "contributor")

# remove_column(open_path, "Metadata", 5)
# remove_row(open_path, "Metadata", 16)
# insert_row(open_saved, "Metadata", 69)
# insert_row(open_saved, "Metadata", 69)
